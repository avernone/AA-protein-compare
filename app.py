import streamlit as st
import requests
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO

st.set_page_config(page_title="Confronto Proteine UniProt", page_icon="🧬", layout="wide")

st.title("🧪 Confronto Amminoacidico tra Proteine (UniProt)")

st.write("""
Inserisci **uno o più codici UniProt (AC)** separati da virgole per confrontare le frequenze relative e i rapporti specifici (E/Q, E/P, Y/F, D/N, G/S) tra più proteine.

Esempio: `P69905, P68871, P02042`
""")

ac_input = st.text_area("Codici UniProt (AC):", "")

if ac_input:
    ac_list = [x.strip() for x in ac_input.split(",") if x.strip()]
    results = {}
    ratio_results = {}
    
    def ratio(a, b, aa_counts):
        return (aa_counts.get(a, 0) / aa_counts.get(b, 1)) if aa_counts.get(b, 0) != 0 else 0
    
    for ac in ac_list:
        url = f"https://rest.uniprot.org/uniprotkb/{ac}.fasta"
        response = requests.get(url)
        
        if response.status_code == 200 and ">" in response.text:
            lines = response.text.splitlines()
            sequence = "".join([l.strip() for l in lines if not l.startswith(">")])
            
            aa_counts = {}
            for aa in sequence:
                aa_counts[aa] = aa_counts.get(aa, 0) + 1
            
            total = len(sequence)
            aa_freq = {aa: count / total for aa, count in aa_counts.items()}
            results[ac] = aa_freq

            ratios = {
                "E/Q": ratio('E', 'Q', aa_counts),
                "E/P": ratio('E', 'P', aa_counts),
                "Y/F": ratio('Y', 'F', aa_counts),
                "D/N": ratio('D', 'N', aa_counts),
                "G/S": ratio('G', 'S', aa_counts)
            }
            ratio_results[ac] = ratios

        else:
            st.warning(f"⚠️ Codice {ac} non valido o sequenza non trovata.")

    if results:
        # Crea DataFrame di confronto per frequenze
        all_aas = sorted({aa for freqs in results.values() for aa in freqs.keys()})
        df_compare = pd.DataFrame(index=all_aas)
        for ac, freqs in results.items():
            df_compare[ac] = [freqs.get(aa, 0) for aa in all_aas]

        st.subheader("📊 Tabella di confronto delle frequenze relative")
        st.dataframe(df_compare.style.format("{:.3f}"))

        # Grafico lineare comparativo
        st.subheader("📈 Grafici comparativi delle frequenze")
        fig, ax = plt.subplots(figsize=(10, 5))
        for ac in ac_list:
            if ac in df_compare.columns:
                ax.plot(df_compare.index, df_compare[ac], label=ac)
        ax.set_xlabel("Aminoacido")
        ax.set_ylabel("Frequenza relativa")
        ax.legend()
        ax.set_title("Confronto frequenze relative tra proteine")
        st.pyplot(fig)

        # Grafico a barre per confronto medio
        st.subheader("📉 Media delle frequenze relative tra le proteine")
        df_compare["Media"] = df_compare.mean(axis=1)
        fig2, ax2 = plt.subplots(figsize=(10, 5))
        ax2.bar(df_compare.index, df_compare["Media"], color="skyblue")
        ax2.set_xlabel("Aminoacido")
        ax2.set_ylabel("Frequenza media relativa")
        ax2.set_title("Distribuzione media amminoacidica tra proteine")
        st.pyplot(fig2)

        # Calcolo rapporti per ogni proteina
        ratio_df = pd.DataFrame(ratio_results).T
        st.subheader("⚖️ Rapporti amminoacidici specifici per proteina")
        st.dataframe(ratio_df.style.format("{:.3f}"))

        # Grafico comparativo dei rapporti
        st.subheader("📊 Confronto rapporti amminoacidici")
        fig3, ax3 = plt.subplots(figsize=(8, 4))
        ratio_df.plot(kind='bar', ax=ax3)
        ax3.set_title("Confronto dei rapporti (E/Q, E/P, Y/F, D/N, G/S)")
        ax3.set_ylabel("Valore del rapporto")
        ax3.set_xlabel("Proteine")
        st.pyplot(fig3)

        # --- 📥 Download risultati in Excel ---
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_compare.to_excel(writer, sheet_name='Confronto_Frequenze')
            ratio_df.to_excel(writer, sheet_name='Confronto_Rapporti')

            img_data = BytesIO()
            fig.savefig(img_data, format='png', bbox_inches='tight')
            img_data.seek(0)
            img_data2 = BytesIO()
            fig2.savefig(img_data2, format='png', bbox_inches='tight')
            img_data2.seek(0)
            img_data3 = BytesIO()
            fig3.savefig(img_data3, format='png', bbox_inches='tight')
            img_data3.seek(0)

            workbook = writer.book
            from openpyxl.drawing.image import Image
            sheet1 = workbook.create_sheet("Grafico_Linee")
            sheet1.add_image(Image(img_data), "A1")
            sheet2 = workbook.create_sheet("Grafico_Barre")
            sheet2.add_image(Image(img_data2), "A1")
            sheet3 = workbook.create_sheet("Grafico_Rapporti")
            sheet3.add_image(Image(img_data3), "A1")

        output.seek(0)

        st.download_button(
            label="📥 Scarica confronto (Excel)",
            data=output,
            file_name="confronto_proteine_uniprot.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

